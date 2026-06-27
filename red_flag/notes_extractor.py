"""
Notes to Accounts Extractor
============================
Extracts Notes to Accounts from annual report PDFs into structured Excel workbooks.

Usage:
    python run.py redflag-notes --companies TCS,INFY --year 2025
"""

import re
import sys
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import ANNUAL_REPORTS_DIR, NOTES_EXCEL_DIR

# ── Styling ──────────────────────────────────────────────────────────────────
COLORS = {
    "tab_header":  "1F4E79",
    "note_header": "2E75B6",
    "alt_row":     "D6E4F0",
    "index_bg":    "F2F2F2",
}
_thin  = Side(style="thin", color="AAAAAA")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

NOTE_HEADING = re.compile(
    r'^\s*(\d{1,2}(?:\.\d+)?)\s{1,6}([A-Z][A-Za-z ,\(\)&\-\/\']{3,80})\s*$'
)
NOTES_KEYWORDS = ["notes to", "note to", "notes forming"]
# Some reports put "Notes" on one line and "to the ... financial statements" on the next.
# We detect this by joining the first few lines and also matching a standalone "Notes" line
# followed by "to the ... financial statements" on the next line.
_NOTES_STANDALONE_RE = re.compile(r'^\s*Notes\s*$', re.IGNORECASE)


def _style_header(cell, bg=COLORS["tab_header"]):
    cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = BORDER


def _style_data(cell, alt=False):
    cell.font      = Font(name="Arial", size=8)
    cell.fill      = PatternFill("solid", fgColor=COLORS["alt_row"] if alt else "FFFFFF")
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.border    = BORDER


def _auto_col_width(ws, max_w=40):
    for col in ws.columns:
        best = 8
        for cell in col:
            if cell.value:
                best = max(best, min(len(str(cell.value)) + 2, max_w))
        ws.column_dimensions[get_column_letter(col[0].column)].width = best


def _find_notes_range(pdf):
    """Return (start_idx, end_idx) 0-based page indices for Notes section.

    Strategy: find pages whose top lines contain a notes-style header.
    Among those, identify distinct section starts (standalone vs consolidated)
    by looking for a gap of 5+ pages between consecutive matches.
    Return the start of the LAST section (consolidated if present, else standalone).
    """
    note_pages = []
    for i, page in enumerate(pdf.pages):
        text = page.extract_text() or ""
        lines = text.split("\n")
        top_text = " ".join(l.strip() for l in lines[:12]).lower()

        # Method 1: keyword appears within the first 12 lines (joined)
        found = any(kw in top_text for kw in NOTES_KEYWORDS)

        # Method 2: "Notes" on its own line followed by "... financial statements"
        if not found:
            for j, line in enumerate(lines[:12]):
                if _NOTES_STANDALONE_RE.match(line) and j + 1 < len(lines):
                    next_line = lines[j + 1].lower().strip()
                    if "financial statement" in next_line:
                        found = True
                        break

        if found:
            note_pages.append(i)

    if not note_pages:
        raise ValueError(
            "Could not find a 'Notes to Financial Statements' section. "
            "The PDF may be scanned/image-based."
        )

    # Group consecutive pages into sections (gap > 5 pages = new section)
    sections = [[note_pages[0]]]
    for p in note_pages[1:]:
        if p - sections[-1][-1] > 5:
            sections.append([p])
        else:
            sections[-1].append(p)

    # Use the last section (consolidated if both exist)
    last_section = sections[-1]
    start = last_section[0]
    end   = last_section[-1]
    return start, end


def _write_table(ws, note_label, note_title, table_idx, table, page_no):
    r      = ws.max_row + 1 if ws.max_row > 1 else 1
    n_cols = max((len(row) for row in table), default=1)
    n_cols = max(n_cols, 2)

    banner = f"Note {note_label}  |  {note_title}  |  Table {table_idx}  |  Page {page_no}"
    ws.cell(row=r, column=1, value=banner)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
    for c in range(1, n_cols + 1):
        _style_header(ws.cell(row=r, column=c), bg=COLORS["note_header"])
    r += 1

    for i, row in enumerate(table):
        for j, val in enumerate(row):
            txt  = (val or "").strip() if val else ""
            cell = ws.cell(row=r, column=j + 1, value=txt)
            _style_data(cell, alt=(i % 2 == 1))
        r += 1

    ws.append([""])


def extract_company(pdf_path, output_path):
    """Extract Notes to Accounts from pdf_path to output_path (.xlsx)."""
    print(f"\n{'\u2500'*60}")
    print(f"  Company : {pdf_path.stem}")
    print(f"  PDF     : {pdf_path.name}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    idx_ws = wb.create_sheet("INDEX")
    idx_ws.sheet_properties.tabColor = "1F4E79"
    for col, (width, header) in enumerate(zip(
        [10, 48, 24, 14, 10],
        ["Note #", "Title", "Sheet Name", "Page(s)", "# Tables"]
    ), 1):
        idx_ws.column_dimensions[get_column_letter(col)].width = width
        _style_header(idx_ws.cell(row=1, column=col))
        idx_ws.cell(row=1, column=col).value = header
    idx_row = 2

    sheet_registry = {}
    table_count    = 0
    current_note   = {"num": "?", "title": "General Notes"}

    with pdfplumber.open(str(pdf_path)) as pdf:
        total = len(pdf.pages)
        print(f"  Pages   : {total}")
        start_idx, end_idx = _find_notes_range(pdf)
        print(f"  Notes   : page {start_idx+1} \u2192 {end_idx+1}  ({end_idx-start_idx+1} pages)")

        for pg_idx in range(start_idx, end_idx + 1):
            page    = pdf.pages[pg_idx]
            page_no = pg_idx + 1
            text    = page.extract_text() or ""
            tables  = page.extract_tables()

            for line in text.split("\n"):
                m = NOTE_HEADING.match(line.strip())
                if m:
                    current_note = {"num": m.group(1), "title": m.group(2).strip()}

            if not tables:
                continue

            raw        = f"N{current_note['num']} - {current_note['title']}"
            sheet_name = re.sub(r'[:\\/?*\[\]]', '', raw)[:31]

            if sheet_name not in sheet_registry:
                ws = wb.create_sheet(sheet_name)
                sheet_registry[sheet_name] = ws
                title_text = f"Note {current_note['num']} \u2014 {current_note['title']}"
                ws.append([title_text])
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
                tc = ws.cell(row=1, column=1)
                tc.font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
                tc.fill      = PatternFill("solid", fgColor=COLORS["tab_header"])
                tc.alignment = Alignment(horizontal="left", vertical="center")
                ws.row_dimensions[1].height = 22
                ws.append([""])
            else:
                ws = sheet_registry[sheet_name]

            for t_idx, table in enumerate(tables, 1):
                if not table or not any(any(c for c in row) for row in table):
                    continue
                table_count += 1
                _write_table(ws, current_note["num"], current_note["title"],
                             t_idx, table, page_no)

                pages_str = f"Pg {page_no}"
                idx_ws.cell(row=idx_row, column=1, value=current_note["num"])
                idx_ws.cell(row=idx_row, column=2, value=current_note["title"])
                idx_ws.cell(row=idx_row, column=3, value=sheet_name)
                idx_ws.cell(row=idx_row, column=4, value=pages_str)
                idx_ws.cell(row=idx_row, column=5, value=t_idx)
                alt = idx_row % 2 == 0
                for c in range(1, 6):
                    cell = idx_ws.cell(row=idx_row, column=c)
                    cell.font   = Font(name="Arial", size=9)
                    cell.fill   = PatternFill(
                        "solid", fgColor=COLORS["index_bg"] if alt else "FFFFFF"
                    )
                    cell.border = BORDER
                idx_row += 1

    for ws in wb.worksheets:
        if ws.title != "INDEX":
            _auto_col_width(ws)
        ws.freeze_panes = ws.cell(row=2, column=1)

    wb.move_sheet("INDEX", offset=-(len(wb.sheetnames) - 1))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    n_sheets = len(wb.sheetnames) - 1
    print(f"  Sheets  : {n_sheets}  |  Tables : {table_count}")
    print(f"  Saved   : {output_path}")

    return {
        "company": pdf_path.stem.split("_")[0],
        "sheets":  n_sheets,
        "tables":  table_count,
        "output":  str(output_path),
    }


def run(companies, year=2025):
    """Extract Notes to Accounts for a list of companies."""
    reports_dir = Path(ANNUAL_REPORTS_DIR)
    output_dir  = Path(NOTES_EXCEL_DIR)

    print(f"\n{'='*60}")
    print(f"  Notes to Accounts Extractor")
    print(f"  Companies : {', '.join(companies)}")
    print(f"  Year      : FY{year}")
    print(f"{'='*60}")

    results = []
    errors  = []

    for ticker in companies:
        pdf_path = reports_dir / ticker / f"{ticker}_AnnualReport_FY{year}.pdf"
        if not pdf_path.exists():
            print(f"\n[SKIP] {ticker}: PDF not found at {pdf_path}")
            errors.append({"company": ticker, "error": "PDF not found"})
            continue

        out_path = output_dir / f"{ticker}_NotesToAccounts_FY{year}.xlsx"

        try:
            summary = extract_company(pdf_path, out_path)
            results.append(summary)
        except Exception as exc:
            print(f"\n[ERROR] {ticker}: {exc}")
            errors.append({"company": ticker, "error": str(exc)})

    print(f"\n{'='*60}")
    print(f"  Extracted : {len(results)} / {len(companies)} companies")
    for r in results:
        print(f"    {r['company']:12s}  {r['sheets']:>3} sheets  {r['tables']:>4} tables")
    if errors:
        print(f"  Errors ({len(errors)}):")
        for e in errors:
            print(f"    {e['company']}: {e['error']}")
    print(f"{'='*60}\n")
