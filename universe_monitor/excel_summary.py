"""Maintain a master Excel summary of all valuations."""

import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from config import OUTPUT_DIR
from shared.data_parser import get_ratio, get_cmp, get_debt_to_equity, get_avg_roce, get_avg_roe, get_pb_ratio
from shared.utils import logger

EXCEL_PATH = os.path.join(OUTPUT_DIR, "valuation_summary.xlsx")
SHEET_NAME = "Valuations"

HEADERS = [
    "Ticker", "Company", "Sector", "CMP (\u20b9)", "Market Cap (Cr)", "P/E",
    "5Y Avg ROCE %", "5Y Avg ROE %", "D/E", "Book Value", "P/B",
    "Weighted IV", "Median IV", "IV Low", "IV High",
    "Upside %", "Methods Agree", "Verdict",
    "P/E vs 5Y Avg", "Value Trap", "Trap Flags",
    "Last Updated",
]

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)


def update_summary(ticker, data, summary):
    """Add or update a row in the master Excel file."""
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    if os.path.exists(EXCEL_PATH):
        wb = load_workbook(EXCEL_PATH)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        _write_headers(ws)

    cmp = get_cmp(data)
    valid_ivs = [r.intrinsic_value for r in summary.results
                 if r.intrinsic_value is not None and r.intrinsic_value > 0]
    iv_low = min(valid_ivs) if valid_ivs else None
    iv_high = max(valid_ivs) if valid_ivs else None

    pe_vs_avg = ""
    if summary.pe_current and summary.pe_5yr_avg:
        pe_vs_avg = f"{summary.pe_current:.1f} vs {summary.pe_5yr_avg:.1f} ({summary.pe_position})"

    row_data = [
        ticker,
        data.get("name", ""),
        summary.sector or "",
        cmp,
        get_ratio(data, "Market Cap"),
        get_ratio(data, "Stock P/E"),
        get_avg_roce(data, n_years=5),
        get_avg_roe(data, n_years=5),
        get_debt_to_equity(data),
        get_ratio(data, "Book Value"),
        get_pb_ratio(data),
        summary.composite_iv,
        summary.median_iv,
        iv_low,
        iv_high,
        summary.upside_pct * 100 if summary.upside_pct is not None else None,
        summary.methods_above_cmp,
        summary.verdict,
        pe_vs_avg,
        summary.value_trap_label,
        "; ".join(summary.value_trap_flags) if summary.value_trap_flags else "",
        datetime.now().strftime("%Y-%m-%d %H:%M"),
    ]

    target_row = None
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == ticker:
            target_row = row_idx
            break

    if target_row is None:
        target_row = ws.max_row + 1

    # Column indices (1-based) for the new header layout:
    # 4=CMP, 10=BV, 11=P/B, 12=Weighted IV, 13=Median IV, 14=IV Low, 15=IV High
    # 16=Upside%, 18=Verdict, 20=Value Trap
    money_cols = {4, 10, 11, 12, 13, 14, 15}
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=target_row, column=col_idx, value=value)
        if col_idx in money_cols:
            cell.number_format = '#,##0.00'
        elif col_idx == 5:  # Market Cap
            cell.number_format = '#,##0'
        elif col_idx == 16:  # Upside %
            cell.number_format = '0.0"%"'

    verdict_cell = ws.cell(row=target_row, column=18)
    if summary.verdict == "BARGAIN":
        verdict_cell.fill = GREEN_FILL
    elif summary.verdict == "OVERVALUED":
        verdict_cell.fill = RED_FILL
    elif summary.verdict in ("UNDERVALUED", "FAIR VALUE"):
        verdict_cell.fill = YELLOW_FILL

    trap_cell = ws.cell(row=target_row, column=20)
    if summary.value_trap_label == "LIKELY TRAP":
        trap_cell.fill = RED_FILL
    elif summary.value_trap_label == "CAUTION":
        trap_cell.fill = ORANGE_FILL
    elif summary.value_trap_label == "CLEAN":
        trap_cell.fill = GREEN_FILL

    from openpyxl.utils import get_column_letter
    for col_idx in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16

    wb.save(EXCEL_PATH)
    logger.info(f"Excel updated for {ticker} \u2014 {summary.verdict}")


def _write_headers(ws):
    """Write formatted header row."""
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{chr(64 + len(HEADERS))}1"
