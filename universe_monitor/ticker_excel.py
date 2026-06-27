"""Generate individual Excel workbook per ticker with full financial data + valuations."""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import TICKER_EXCELS_DIR
from shared.data_parser import get_ratio, get_cmp
from shared.utils import logger

# Styles
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SECTION_FONT = Font(bold=True, size=11)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
GREEN_FONT = Font(color="006100")
RED_FONT = Font(color="9C0006")
BOLD = Font(bold=True)
THIN_BORDER = Border(
    bottom=Side(style="thin", color="CCCCCC"),
)
NUM_FMT = '#,##0.00'
NUM_FMT_INT = '#,##0'
PCT_FMT = '0.0"%"'


def write_ticker_excel(ticker, data, summary):
    """Create a comprehensive Excel workbook for a single ticker."""
    os.makedirs(TICKER_EXCELS_DIR, exist_ok=True)
    filepath = os.path.join(TICKER_EXCELS_DIR, f"{ticker}.xlsx")

    wb = Workbook()

    _write_snapshot_sheet(wb, data, summary)
    _write_financial_sheet(wb, data, "profit_loss", "Profit & Loss")
    _write_financial_sheet(wb, data, "balance_sheet", "Balance Sheet")
    _write_financial_sheet(wb, data, "cash_flow", "Cash Flow")
    _write_financial_sheet(wb, data, "ratios", "Ratios")
    _write_financial_sheet(wb, data, "quarters", "Quarterly")
    _write_financial_sheet(wb, data, "shareholding", "Shareholding")

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    wb.save(filepath)
    logger.info(f"Ticker Excel written: {filepath}")
    return filepath


def _write_snapshot_sheet(wb, data, summary):
    """Sheet 1: Company snapshot + all 10 valuation results."""
    ws = wb.active
    ws.title = "Valuation"

    name = data.get("name", data.get("ticker", ""))
    ticker = data.get("ticker", "")
    cmp = get_cmp(data)
    row = 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=1, value=f"{name} ({ticker})")
    cell.font = Font(bold=True, size=14, color="4472C4")
    row += 1

    ws.cell(row=row, column=1, value=f"Analysis Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws.cell(row=row, column=1).font = Font(italic=True, color="808080")
    row += 2

    _section_header(ws, row, "Key Ratios", cols=4)
    row += 1

    ratio_items = [
        ("Current Price", "Current Price"),
        ("Market Cap (Cr)", "Market Cap"),
        ("Stock P/E", "Stock P/E"),
        ("Book Value", "Book Value"),
        ("ROCE %", "ROCE"),
        ("ROE %", "ROE"),
        ("Debt to Equity", "Debt to equity"),
        ("Dividend Yield %", "Dividend Yield"),
        ("Promoter Holding %", "Promoter holding"),
        ("Industry P/E", "Industry PE"),
        ("High / Low", "High / Low"),
        ("Face Value", "Face Value"),
    ]

    for label, key in ratio_items:
        val = data.get("top_ratios", {}).get(key, "N/A")
        ws.cell(row=row, column=1, value=label).font = BOLD
        ws.cell(row=row, column=2, value=val)
        row += 1

    row += 1

    _section_header(ws, row, "Valuation Summary", cols=6)
    row += 1

    ws.cell(row=row, column=1, value="Verdict").font = BOLD
    verdict_cell = ws.cell(row=row, column=2, value=summary.verdict)
    verdict_cell.font = Font(bold=True, size=12)
    if summary.verdict == "BARGAIN":
        verdict_cell.fill = GREEN_FILL
        verdict_cell.font = Font(bold=True, size=12, color="006100")
    elif summary.verdict == "OVERVALUED":
        verdict_cell.fill = RED_FILL
        verdict_cell.font = Font(bold=True, size=12, color="9C0006")
    elif summary.verdict in ("UNDERVALUED", "FAIR VALUE"):
        verdict_cell.fill = YELLOW_FILL
    row += 1

    ws.cell(row=row, column=1, value="Sector").font = BOLD
    ws.cell(row=row, column=2, value=summary.sector or "N/A")
    row += 1

    ws.cell(row=row, column=1, value="Weighted Intrinsic Value").font = BOLD
    if summary.composite_iv:
        ws.cell(row=row, column=2, value=summary.composite_iv).number_format = NUM_FMT
    else:
        ws.cell(row=row, column=2, value="N/A")
    row += 1

    ws.cell(row=row, column=1, value="Median IV (reference)").font = BOLD
    if summary.median_iv:
        ws.cell(row=row, column=2, value=summary.median_iv).number_format = NUM_FMT
    else:
        ws.cell(row=row, column=2, value="N/A")
    row += 1

    ws.cell(row=row, column=1, value="Upside / Downside").font = BOLD
    if summary.upside_pct is not None:
        pct_cell = ws.cell(row=row, column=2, value=f"{summary.upside_pct * 100:+.1f}%")
        if summary.upside_pct > 0:
            pct_cell.font = GREEN_FONT
        else:
            pct_cell.font = RED_FONT
    row += 1

    ws.cell(row=row, column=1, value="Methods Above CMP").font = BOLD
    ws.cell(row=row, column=2, value=f"{summary.methods_above_cmp} of {len(summary.results)}")
    row += 1

    ws.cell(row=row, column=1, value="Value Trap").font = BOLD
    trap_cell = ws.cell(row=row, column=2, value=summary.value_trap_label)
    if summary.value_trap_label == "LIKELY TRAP":
        trap_cell.fill = RED_FILL
        trap_cell.font = Font(bold=True, color="9C0006")
    elif summary.value_trap_label == "CAUTION":
        trap_cell.fill = YELLOW_FILL
    elif summary.value_trap_label == "CLEAN":
        trap_cell.fill = GREEN_FILL
        trap_cell.font = Font(bold=True, color="006100")
    row += 1

    if summary.value_trap_flags:
        ws.cell(row=row, column=1, value="Trap Flags").font = BOLD
        ws.cell(row=row, column=2, value="; ".join(summary.value_trap_flags))
        row += 1

    row += 1

    _section_header(ws, row, "Valuation Methods Detail", cols=6)
    row += 1

    headers = ["#", "Method", "Intrinsic Value (\u20b9)", "vs CMP", "Confidence", "Notes"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    row += 1

    for i, r in enumerate(summary.results, 1):
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=r.method).font = BOLD

        if r.intrinsic_value:
            ws.cell(row=row, column=3, value=r.intrinsic_value).number_format = NUM_FMT
        else:
            ws.cell(row=row, column=3, value="N/A")

        if r.intrinsic_value and cmp and cmp > 0:
            diff = ((r.intrinsic_value / cmp) - 1) * 100
            diff_cell = ws.cell(row=row, column=4, value=f"{diff:+.1f}%")
            if diff > 0:
                diff_cell.font = GREEN_FONT
            else:
                diff_cell.font = RED_FONT
        else:
            ws.cell(row=row, column=4, value="-")

        ws.cell(row=row, column=5, value=r.confidence)
        ws.cell(row=row, column=6, value=r.notes)

        if i % 2 == 0:
            for c in range(1, 7):
                ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    col_widths = [5, 25, 20, 12, 12, 60]
    for idx, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def _write_financial_sheet(wb, data, section_key, sheet_name):
    """Write a financial data table as a sheet."""
    section = data.get(section_key, {})
    years = section.get("years", [])
    metrics = section.get("data", {})

    if not years or not metrics:
        return

    ws = wb.create_sheet(title=sheet_name)

    ws.cell(row=1, column=1, value="Metric").fill = HEADER_FILL
    ws.cell(row=1, column=1).font = HEADER_FONT
    for col_idx, year in enumerate(years, 2):
        cell = ws.cell(row=1, column=col_idx, value=year)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    row = 2
    for metric_name, values in metrics.items():
        ws.cell(row=row, column=1, value=metric_name).font = BOLD

        for col_idx, val in enumerate(values, 2):
            if val is not None:
                cell = ws.cell(row=row, column=col_idx, value=val)
                if "%" in metric_name or "OPM" in metric_name or "margin" in metric_name.lower():
                    cell.number_format = '0.00"%"'
                else:
                    cell.number_format = NUM_FMT
            else:
                ws.cell(row=row, column=col_idx, value="")

        if row % 2 == 0:
            for c in range(1, len(years) + 2):
                ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    ws.column_dimensions["A"].width = 30
    for col_idx in range(2, len(years) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    ws.freeze_panes = "B2"


def _section_header(ws, row, text, cols=6):
    """Write a section header spanning multiple columns."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = SECTION_FILL
    cell.font = SECTION_FONT
