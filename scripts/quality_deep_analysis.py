"""
Quality Business Deep Analysis — Forward-Looking Report
Analyzes all companies passing the quality screen with:
- 5 Pointer scores
- Incremental ROCE calculation
- Reinvestment capacity
- 5-year and 10-year EPS/stock price projections (bear/base/bull)
"""
import sys, io, os, warnings, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
warnings.filterwarnings('ignore')
import openpyxl
import statistics
from datetime import datetime

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "ticker_excels")
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "output", "quality_reports")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# The 43 tickers that passed the quality screen
QUALITY_TICKERS = [
    "TRITURBINE", "CAMS", "KFINTECH", "KPITTECH", "PERSISTENT", "TATATECH",
    "AGIIL", "BLS", "CUMMINSIND", "ECLERX", "EICHERMOT", "IEX",
    "INDIAMART", "LALPATHLAB", "PIIND", "ALKYLAMINE", "CDSL", "CRISIL",
    "GRWRHITECH", "JBCHEPHARM", "MCX", "SOLARINDS", "SUPREMEIND", "TRAVELFOOD",
    "CONTROLPR", "GRAUWEIL", "GRINDWELL", "INOXINDIA", "JLHL", "LTM",
    "MEDANTA", "VIJAYA", "ALLDIGI", "JYOTHYLAB", "NATIONALUM", "PIDILITIND",
    "RAINBOW", "TATAELXSI", "ZYDUSLIFE", "COFORGE", "NH", "GPIL", "INDUSTOWER"
]


def safe_float(v):
    if v is None or v == 'N/A' or v == '' or v == '-':
        return None
    try:
        if isinstance(v, str):
            v = v.replace(',', '').replace('%', '').strip()
        return float(v)
    except:
        return None


def get_val(ws, key):
    for row in ws.iter_rows(values_only=True):
        if row[0] and str(row[0]).strip() == key:
            return row[1]
    return None


def get_sheet_data(wb, sheet_name):
    ws = wb[sheet_name]
    data = {}
    headers = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = list(row)
        elif row[0]:
            data[str(row[0]).strip()] = list(row[1:])
    return headers, data


def cagr(start, end, years):
    if not start or start <= 0 or not end or end <= 0 or years <= 0:
        return None
    return ((end / start) ** (1 / years) - 1) * 100


def analyze_company(ticker):
    fpath = os.path.join(DATA_DIR, f"{ticker}.xlsx")
    if not os.path.exists(fpath):
        return None

    try:
        wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
    except:
        return None

    result = {'ticker': ticker}

    # ===== VALUATION SHEET =====
    ws_val = wb['Valuation']
    result['name'] = get_val(ws_val, None) or ticker  # First cell has company name
    # Get name from first row
    for row in ws_val.iter_rows(min_row=1, max_row=1, values_only=True):
        result['name'] = str(row[0]).split('(')[0].strip() if row[0] else ticker

    result['cmp'] = safe_float(get_val(ws_val, 'Current Price'))
    result['mcap'] = safe_float(get_val(ws_val, 'Market Cap (Cr)'))
    result['pe'] = safe_float(get_val(ws_val, 'Stock P/E'))
    result['bv'] = safe_float(get_val(ws_val, 'Book Value'))
    result['roce'] = safe_float(get_val(ws_val, 'ROCE %'))
    result['roe'] = safe_float(get_val(ws_val, 'ROE %'))
    result['div_yield'] = safe_float(get_val(ws_val, 'Dividend Yield %'))
    result['sector'] = get_val(ws_val, 'Sector') or '-'
    result['face_value'] = safe_float(get_val(ws_val, 'Face Value'))

    # ===== P&L =====
    headers_pl, pl = get_sheet_data(wb, 'Profit & Loss')
    year_headers = [str(h) for h in headers_pl[1:] if h and 'TTM' not in str(h)]
    n = len(year_headers)

    sales_raw = [safe_float(x) for x in pl.get('Sales', [])[:n]]
    op_raw = [safe_float(x) for x in pl.get('Operating Profit', [])[:n]]
    opm_raw = [safe_float(x) for x in pl.get('OPM %', [])[:n]]
    np_raw = [safe_float(x) for x in pl.get('Net Profit', [])[:n]]
    other_income = [safe_float(x) for x in pl.get('Other Income', [])[:n]]
    interest = [safe_float(x) for x in pl.get('Interest', [])[:n]]
    depreciation = [safe_float(x) for x in pl.get('Depreciation', [])[:n]]
    tax_pct = [safe_float(x) for x in pl.get('Tax %', [])[:n]]

    # Clean arrays (last N valid values)
    sales = [s for s in sales_raw if s is not None]
    op = [o for o in op_raw if o is not None]
    opm = [o for o in opm_raw if o is not None]
    np_list = [n for n in np_raw if n is not None]
    depr = [d for d in (depreciation or []) if d is not None]

    result['revenue_latest'] = sales[-1] if sales else None
    result['revenue_5y_ago'] = sales[-6] if len(sales) >= 6 else (sales[0] if sales else None)
    result['pat_latest'] = np_list[-1] if np_list else None
    result['pat_5y_ago'] = np_list[-6] if len(np_list) >= 6 else (np_list[0] if np_list else None)
    result['op_latest'] = op[-1] if op else None

    result['sales_cagr_5y'] = cagr(result['revenue_5y_ago'], result['revenue_latest'], 5)
    result['sales_cagr_3y'] = cagr(sales[-4] if len(sales) >= 4 else None, sales[-1] if sales else None, 3)
    result['sales_cagr_10y'] = cagr(sales[0] if len(sales) >= 10 else None, sales[-1] if sales else None, min(n-1, 10))
    result['pat_cagr_5y'] = cagr(result['pat_5y_ago'], result['pat_latest'], 5)
    result['pat_cagr_3y'] = cagr(np_list[-4] if len(np_list) >= 4 else None, np_list[-1] if np_list else None, 3)

    # OPM analysis
    opm_last5 = [o for o in opm[-5:] if o is not None]
    result['opm_avg'] = statistics.mean(opm_last5) if len(opm_last5) >= 3 else None
    result['opm_range'] = (max(opm_last5) - min(opm_last5)) if len(opm_last5) >= 3 else None
    result['opm_latest'] = opm[-1] if opm else None
    result['opm_trend'] = 'Expanding' if len(opm) >= 3 and opm[-1] > opm[-3] else ('Stable' if len(opm) >= 3 and abs(opm[-1] - opm[-3]) <= 3 else 'Compressing')

    # Revenue acceleration
    if result['sales_cagr_3y'] and result['sales_cagr_5y']:
        result['rev_accelerating'] = result['sales_cagr_3y'] >= result['sales_cagr_5y'] * 0.8
    else:
        result['rev_accelerating'] = None

    # Operating leverage
    result['op_leverage'] = (result['pat_cagr_5y'] or 0) > (result['sales_cagr_5y'] or 0)

    # Tax stability
    tax_last5 = [t for t in (tax_pct or [])[-5:] if t is not None]
    result['tax_stable'] = (max(tax_last5) - min(tax_last5)) <= 10 if len(tax_last5) >= 3 else None

    # Depreciation (latest)
    result['depreciation'] = depr[-1] if depr else 0

    # ===== BALANCE SHEET =====
    _, bs = get_sheet_data(wb, 'Balance Sheet')
    borrowings = [safe_float(x) for x in bs.get('Borrowings', [])]
    reserves = [safe_float(x) for x in bs.get('Reserves', [])]
    equity_cap = [safe_float(x) for x in bs.get('Equity Capital', [])]
    other_liab = [safe_float(x) for x in bs.get('Other Liabilities', [])]
    fixed_assets = [safe_float(x) for x in bs.get('Fixed Assets', [])]
    cwip = [safe_float(x) for x in bs.get('CWIP', [])]
    investments = [safe_float(x) for x in bs.get('Investments', [])]
    total_assets_row = [safe_float(x) for x in bs.get('Total Assets', [])]

    borr_clean = [b for b in borrowings if b is not None]
    res_clean = [r for r in reserves if r is not None]
    eq_clean = [e for e in equity_cap if e is not None]

    result['borrowings'] = borr_clean[-1] if borr_clean else 0
    result['reserves'] = res_clean[-1] if res_clean else 0
    result['equity_capital'] = eq_clean[-1] if eq_clean else 0
    result['total_equity'] = (result['reserves'] or 0) + (result['equity_capital'] or 0)
    result['de_ratio'] = (result['borrowings'] or 0) / result['total_equity'] if result['total_equity'] > 0 else 0
    result['fixed_assets'] = ([f for f in fixed_assets if f is not None] or [0])[-1]
    result['cwip'] = ([c for c in cwip if c is not None] or [0])[-1]

    # Capital employed (latest and 4 years ago for incremental ROCE)
    def calc_capital_employed(idx):
        eq = eq_clean[idx] if idx < len(eq_clean) else None
        res = res_clean[idx] if idx < len(res_clean) else None
        borr = borr_clean[idx] if idx < len(borr_clean) else None
        if eq is not None and res is not None:
            return (eq or 0) + (res or 0) + (borr or 0)
        return None

    ce_latest = calc_capital_employed(-1)
    ce_4y = calc_capital_employed(-5) if len(eq_clean) >= 5 else calc_capital_employed(0)

    # EBIT (PBT + Interest)
    int_clean = [safe_float(x) for x in (interest or []) if safe_float(x) is not None]
    pbt_list = []
    for i in range(n):
        s = sales_raw[i] if i < len(sales_raw) else None
        e_val = safe_float(pl.get('Expenses', [None]*n)[i]) if i < n else None
        oi = other_income[i] if other_income and i < len(other_income) else None
        intr = interest[i] if interest and i < len(interest) else None
        dep = depreciation[i] if depreciation and i < len(depreciation) else None
        # EBIT = OP + Other Income
        if op_raw[i] is not None and (other_income and i < len(other_income)):
            oi_val = other_income[i] or 0
            intr_val = interest[i] if interest and i < len(interest) else 0
            ebit = (op_raw[i] or 0) + (oi_val or 0)
            pbt_list.append(ebit)
        elif op_raw[i] is not None:
            pbt_list.append(op_raw[i])

    ebit_latest = pbt_list[-1] if pbt_list else None
    ebit_4y = pbt_list[-5] if len(pbt_list) >= 5 else (pbt_list[0] if pbt_list else None)

    # Incremental ROCE
    if ce_latest and ce_4y and ebit_latest and ebit_4y and (ce_latest - ce_4y) > 0:
        result['incremental_roce'] = ((ebit_latest - ebit_4y) / (ce_latest - ce_4y)) * 100
    else:
        result['incremental_roce'] = None

    result['capital_employed'] = ce_latest

    # ===== CASH FLOW =====
    _, cf = get_sheet_data(wb, 'Cash Flow')
    cfo_raw = [safe_float(x) for x in cf.get('Cash from Operating Activity', [])]
    cfi_raw = [safe_float(x) for x in cf.get('Cash from Investing Activity', [])]
    cff_raw = [safe_float(x) for x in cf.get('Cash from Financing Activity', [])]

    cfo = [c for c in cfo_raw if c is not None]
    result['cfo_latest'] = cfo[-1] if cfo else None

    # CFO/OP ratios
    cfo_last5 = cfo[-5:]
    op_last5 = op[-5:]
    min_len = min(len(cfo_last5), len(op_last5))
    cfo_op_ratios = []
    for i in range(min_len):
        if op_last5[i] and op_last5[i] > 0:
            cfo_op_ratios.append(cfo_last5[i] / op_last5[i] * 100)
    result['avg_cfo_op'] = statistics.mean(cfo_op_ratios) if cfo_op_ratios else None
    result['cfo_positive_streak'] = sum(1 for c in cfo[-5:] if c and c > 0)

    # ===== RATIOS =====
    _, rat = get_sheet_data(wb, 'Ratios')
    dd = [safe_float(x) for x in rat.get('Debtor Days', [])]
    inv_days = [safe_float(x) for x in rat.get('Inventory Days', [])]
    wc_days = [safe_float(x) for x in rat.get('Cash Conversion Cycle', [])]

    dd_clean = [d for d in dd if d is not None]
    result['debtor_days'] = dd_clean[-1] if dd_clean else None
    wc_clean = [w for w in wc_days if w is not None]
    result['wc_days'] = wc_clean[-1] if wc_clean else None

    # WC trend
    if len(wc_clean) >= 3:
        result['wc_trend'] = 'Improving' if wc_clean[-1] < wc_clean[-3] else ('Stable' if abs(wc_clean[-1] - wc_clean[-3]) < 15 else 'Deteriorating')
    else:
        result['wc_trend'] = None

    # ===== SHAREHOLDING =====
    _, sh = get_sheet_data(wb, 'Shareholding')
    prom = [safe_float(x) for x in sh.get('Promoters', [])]
    fiis = [safe_float(x) for x in sh.get('FIIs', [])]
    diis = [safe_float(x) for x in sh.get('DIIs', [])]

    prom_clean = [p for p in prom if p is not None]
    fii_clean = [f for f in fiis if f is not None]
    dii_clean = [d for d in diis if d is not None]

    result['promoter_latest'] = prom_clean[-1] if prom_clean else 0
    result['promoter_change'] = (prom_clean[-1] - prom_clean[0]) if len(prom_clean) >= 2 else 0
    result['fii_latest'] = fii_clean[-1] if fii_clean else 0
    result['dii_latest'] = dii_clean[-1] if dii_clean else 0

    # ===== FORWARD PROJECTIONS =====
    # EPS calculation
    if result['cmp'] and result['pe'] and result['pe'] > 0:
        result['eps_current'] = result['cmp'] / result['pe']
    elif result['pat_latest'] and result['mcap'] and result['cmp']:
        shares = result['mcap'] / result['cmp']  # in crores
        result['eps_current'] = result['pat_latest'] / shares if shares > 0 else None
    else:
        result['eps_current'] = None

    # Reinvestment capacity
    retained_pct = 1 - ((result['div_yield'] or 0) / 100 * (result['pe'] or 20) / 100)  # rough
    retained_earnings = (result['pat_latest'] or 0) * max(0.7, min(retained_pct, 0.98))
    depr_val = result.get('depreciation', 0) or 0
    result['internal_accruals'] = retained_earnings + depr_val

    # Use incremental ROCE if available, else use current ROCE
    inc_roce = result.get('incremental_roce')
    curr_roce = result.get('roce', 15)

    # Projected earnings growth rate
    # Method: min of (reinvestment-implied growth, historical PAT CAGR) for conservatism
    if inc_roce and inc_roce > 0 and result['capital_employed'] and result['capital_employed'] > 0:
        reinvestment_rate = result['internal_accruals'] / result['capital_employed']
        implied_growth = reinvestment_rate * (inc_roce / 100) * 100
    else:
        implied_growth = (result['pat_cagr_5y'] or 12) * 0.8

    historical_growth = result['pat_cagr_5y'] or 12

    # Base case: average of implied and historical, capped
    base_growth = min(max((implied_growth + historical_growth) / 2, 8), 30)
    bear_growth = max(base_growth * 0.6, 5)
    bull_growth = min(base_growth * 1.3, 35)

    result['growth_bear'] = round(bear_growth, 1)
    result['growth_base'] = round(base_growth, 1)
    result['growth_bull'] = round(bull_growth, 1)

    # 5-year projections
    eps = result['eps_current']
    if eps and eps > 0:
        result['eps_5y_bear'] = round(eps * (1 + bear_growth/100)**5, 1)
        result['eps_5y_base'] = round(eps * (1 + base_growth/100)**5, 1)
        result['eps_5y_bull'] = round(eps * (1 + bull_growth/100)**5, 1)

        result['eps_10y_bear'] = round(eps * (1 + bear_growth/100)**10, 1)
        result['eps_10y_base'] = round(eps * (1 + base_growth/100)**10, 1)
        result['eps_10y_bull'] = round(eps * (1 + bull_growth/100)**10, 1)

        # PE assumptions for future
        curr_pe = result['pe'] or 20
        result['pe_bear'] = round(max(curr_pe * 0.7, 12), 1)
        result['pe_base'] = round(curr_pe, 1)
        result['pe_bull'] = round(min(curr_pe * 1.3, 60), 1)

        result['price_5y_bear'] = round(result['eps_5y_bear'] * result['pe_bear'])
        result['price_5y_base'] = round(result['eps_5y_base'] * result['pe_base'])
        result['price_5y_bull'] = round(result['eps_5y_bull'] * result['pe_bull'])

        result['price_10y_bear'] = round(result['eps_10y_bear'] * result['pe_bear'])
        result['price_10y_base'] = round(result['eps_10y_base'] * result['pe_base'])
        result['price_10y_bull'] = round(result['eps_10y_bull'] * result['pe_bull'])

        # CAGR from CMP
        cmp = result['cmp']
        if cmp and cmp > 0:
            result['cagr_5y_bear'] = round(cagr(cmp, result['price_5y_bear'], 5) or 0, 1)
            result['cagr_5y_base'] = round(cagr(cmp, result['price_5y_base'], 5) or 0, 1)
            result['cagr_5y_bull'] = round(cagr(cmp, result['price_5y_bull'], 5) or 0, 1)
            result['cagr_10y_bear'] = round(cagr(cmp, result['price_10y_bear'], 10) or 0, 1)
            result['cagr_10y_base'] = round(cagr(cmp, result['price_10y_base'], 10) or 0, 1)
            result['cagr_10y_bull'] = round(cagr(cmp, result['price_10y_bull'], 10) or 0, 1)

    # ===== QUALITY SCORE =====
    score = 0
    reasons = []

    # OPM stability
    if result['opm_range'] is not None:
        if result['opm_range'] <= 5:
            score += 3; reasons.append("Very stable margins (<=5% range)")
        elif result['opm_range'] <= 8:
            score += 2; reasons.append("Stable margins (<=8% range)")
        elif result['opm_range'] <= 12:
            score += 1; reasons.append("Moderate margin stability")

    # CFO/OP
    if result['avg_cfo_op'] is not None:
        if result['avg_cfo_op'] >= 100:
            score += 3; reasons.append("Exceptional cash conversion (CFO/OP >100%)")
        elif result['avg_cfo_op'] >= 85:
            score += 2; reasons.append("Strong cash conversion (CFO/OP >85%)")
        elif result['avg_cfo_op'] >= 70:
            score += 1; reasons.append("Adequate cash conversion")

    # Growth
    if (result['sales_cagr_5y'] or 0) >= 18:
        score += 2; reasons.append("High revenue growth (>18% CAGR)")
    elif (result['sales_cagr_5y'] or 0) >= 12:
        score += 1; reasons.append("Good revenue growth (>12% CAGR)")

    if (result['pat_cagr_5y'] or 0) >= 20:
        score += 2; reasons.append("High profit growth (>20% CAGR)")
    elif (result['pat_cagr_5y'] or 0) >= 15:
        score += 1; reasons.append("Good profit growth (>15% CAGR)")

    # ROCE
    if (result['roce'] or 0) >= 25:
        score += 2; reasons.append("Excellent ROCE (>25%)")
    elif (result['roce'] or 0) >= 20:
        score += 1; reasons.append("Good ROCE (>20%)")

    if result.get('rev_accelerating'): score += 1; reasons.append("Revenue accelerating")
    if result.get('op_leverage'): score += 1; reasons.append("Operating leverage present")
    if (result.get('promoter_change') or 0) > 0: score += 1; reasons.append("Promoter buying")
    if (result.get('de_ratio') or 0) < 0.1: score += 1; reasons.append("Virtually debt-free")
    if (result.get('debtor_days') or 999) < 30: score += 1; reasons.append("Low debtor days (<30)")

    result['quality_score'] = score
    result['quality_reasons'] = reasons

    # Verdict
    if score >= 13:
        result['verdict'] = 'HIGHEST QUALITY'
    elif score >= 11:
        result['verdict'] = 'HIGH QUALITY'
    elif score >= 9:
        result['verdict'] = 'QUALITY'
    else:
        result['verdict'] = 'WATCHLIST'

    wb.close()
    return result


def generate_report(r):
    """Generate markdown report for a company"""
    lines = []
    lines.append(f"# {r['name']} ({r['ticker']})")
    lines.append(f"**Sector:** {r['sector']} | **Verdict:** {r['verdict']} | **Quality Score:** {r['quality_score']}/18")
    lines.append("")
    lines.append("---")
    lines.append("")

    # Key metrics table
    lines.append("## Key Metrics")
    lines.append("")
    lines.append("| Metric | Value |")
    lines.append("|--------|-------|")
    lines.append(f"| CMP | Rs {r.get('cmp', '-')} |")
    lines.append(f"| Market Cap | Rs {r.get('mcap', '-')} Cr |")
    lines.append(f"| PE | {r.get('pe', '-')}x |")
    lines.append(f"| Book Value | Rs {r.get('bv', '-')} |")
    lines.append(f"| ROCE | {r.get('roce', '-')}% |")
    lines.append(f"| ROE | {r.get('roe', '-')}% |")
    lines.append(f"| D/E Ratio | {r.get('de_ratio', '-'):.2f} |" if r.get('de_ratio') is not None else "| D/E Ratio | - |")
    lines.append(f"| Dividend Yield | {r.get('div_yield', '-')}% |")
    lines.append(f"| Promoter Holding | {r.get('promoter_latest', '-')}% (chg: {r.get('promoter_change', 0):+.1f}pp) |")
    lines.append("")

    # Growth
    lines.append("## Growth Profile")
    lines.append("")
    lines.append("| Metric | 10Y | 5Y | 3Y |")
    lines.append("|--------|-----|-----|-----|")
    s10 = f"{r['sales_cagr_10y']:.1f}" if r.get('sales_cagr_10y') is not None else "-"
    s5 = f"{r['sales_cagr_5y']:.1f}" if r.get('sales_cagr_5y') is not None else "-"
    s3 = f"{r['sales_cagr_3y']:.1f}" if r.get('sales_cagr_3y') is not None else "-"
    lines.append(f"| Revenue CAGR | {s10}% | {s5}% | {s3}% |")
    p5 = f"{r['pat_cagr_5y']:.1f}" if r.get('pat_cagr_5y') is not None else "-"
    p3 = f"{r['pat_cagr_3y']:.1f}" if r.get('pat_cagr_3y') is not None else "-"
    lines.append(f"| PAT CAGR | - | {p5}% | {p3}% |")
    lines.append("")
    lines.append(f"- Revenue trend: **{'Accelerating' if r.get('rev_accelerating') else 'Decelerating'}**")
    lines.append(f"- Operating leverage: **{'Yes' if r.get('op_leverage') else 'No'}** (PAT growing {'faster' if r.get('op_leverage') else 'slower'} than revenue)")
    lines.append("")

    # Quality signals
    lines.append("## Quality Signals")
    lines.append("")
    lines.append("| Signal | Value | Assessment |")
    lines.append("|--------|-------|------------|")

    opm_assess = "Excellent" if (r.get('opm_range') or 99) <= 5 else ("Good" if (r.get('opm_range') or 99) <= 8 else "Moderate")
    lines.append(f"| OPM (5Y avg) | {r.get('opm_avg', '-'):.1f}% (range: {r.get('opm_range', '-'):.1f}%) | {opm_assess} |" if r.get('opm_avg') else "| OPM | - | - |")

    cfo_assess = "Exceptional" if (r.get('avg_cfo_op') or 0) >= 100 else ("Strong" if (r.get('avg_cfo_op') or 0) >= 85 else "Adequate")
    lines.append(f"| CFO/OP (5Y avg) | {r.get('avg_cfo_op', '-'):.1f}% | {cfo_assess} |" if r.get('avg_cfo_op') else "| CFO/OP | - | - |")

    lines.append(f"| Debtor Days | {r.get('debtor_days', '-')} | {'Clean' if (r.get('debtor_days') or 999) < 30 else ('OK' if (r.get('debtor_days') or 999) < 60 else 'Watch')} |")
    lines.append(f"| Working Capital | {r.get('wc_days', '-')} days ({r.get('wc_trend', '-')}) | {'Good' if r.get('wc_trend') != 'Deteriorating' else 'Concern'} |")
    lines.append(f"| Tax Rate | {'Stable' if r.get('tax_stable') else 'Variable'} | {'Clean' if r.get('tax_stable') else 'Check'} |")
    lines.append(f"| Debt | Rs {r.get('borrowings', 0):.0f} Cr (D/E: {r.get('de_ratio', 0):.2f}) | {'Clean' if r.get('de_ratio', 1) < 0.3 else 'Moderate'} |")
    lines.append("")

    # Quality reasons
    lines.append("**Quality drivers:** " + " | ".join(r.get('quality_reasons', [])))
    lines.append("")

    # Forward projections
    lines.append("## Forward-Looking Projections")
    lines.append("")

    if r.get('incremental_roce'):
        lines.append(f"**Incremental ROCE:** {r['incremental_roce']:.1f}% (return on each new rupee invested)")
    lines.append(f"**Internal accruals:** Rs {r.get('internal_accruals', 0):.0f} Cr/year (retained earnings + depreciation)")
    lines.append(f"**Current EPS:** Rs {r.get('eps_current', 0):.1f}")
    lines.append("")

    if r.get('eps_5y_bear'):
        lines.append("### 5-Year Projection")
        lines.append("")
        lines.append("| Scenario | EPS CAGR | FY31 EPS | PE | Stock Price | CAGR from CMP |")
        lines.append("|----------|----------|----------|----|-------------|---------------|")
        lines.append(f"| Bear | {r['growth_bear']}% | Rs {r['eps_5y_bear']} | {r['pe_bear']}x | Rs {r['price_5y_bear']:,} | {r.get('cagr_5y_bear', '-')}% |")
        lines.append(f"| Base | {r['growth_base']}% | Rs {r['eps_5y_base']} | {r['pe_base']}x | Rs {r['price_5y_base']:,} | {r.get('cagr_5y_base', '-')}% |")
        lines.append(f"| Bull | {r['growth_bull']}% | Rs {r['eps_5y_bull']} | {r['pe_bull']}x | Rs {r['price_5y_bull']:,} | {r.get('cagr_5y_bull', '-')}% |")
        lines.append("")

        lines.append("### 10-Year Projection")
        lines.append("")
        lines.append("| Scenario | EPS CAGR | FY36 EPS | PE | Stock Price | CAGR from CMP |")
        lines.append("|----------|----------|----------|----|-------------|---------------|")
        lines.append(f"| Bear | {r['growth_bear']}% | Rs {r['eps_10y_bear']} | {r['pe_bear']}x | Rs {r['price_10y_bear']:,} | {r.get('cagr_10y_bear', '-')}% |")
        lines.append(f"| Base | {r['growth_base']}% | Rs {r['eps_10y_base']} | {r['pe_base']}x | Rs {r['price_10y_base']:,} | {r.get('cagr_10y_base', '-')}% |")
        lines.append(f"| Bull | {r['growth_bull']}% | Rs {r['eps_10y_bull']} | {r['pe_bull']}x | Rs {r['price_10y_bull']:,} | {r.get('cagr_10y_bull', '-')}% |")
        lines.append("")

    lines.append("---")
    lines.append(f"*Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}*")

    return "\n".join(lines)


# ===== MAIN =====
print(f"Analyzing {len(QUALITY_TICKERS)} quality companies...\n")

all_results = []
for ticker in QUALITY_TICKERS:
    r = analyze_company(ticker)
    if r:
        all_results.append(r)
        # Save individual report
        report = generate_report(r)
        with open(os.path.join(OUTPUT_DIR, f"{ticker}.md"), 'w', encoding='utf-8') as f:
            f.write(report)
        print(f"  OK  {ticker:<16} Score:{r['quality_score']:>3}  ROCE:{r.get('roce',0):>5.1f}%  5Y-Base CAGR:{r.get('cagr_5y_base', 0):>6.1f}%  10Y-Base CAGR:{r.get('cagr_10y_base', 0):>6.1f}%")
    else:
        print(f"  SKIP {ticker}")

# Sort by quality score, then by base case 10Y CAGR
all_results.sort(key=lambda x: (x['quality_score'], x.get('cagr_10y_base', 0)), reverse=True)

# Generate master summary
summary_lines = []
summary_lines.append("# Quality Business Universe — Master Report")
summary_lines.append(f"*Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Companies: {len(all_results)}*")
summary_lines.append("")
summary_lines.append("## Screening Criteria Applied")
summary_lines.append("- ROCE > 18%, ROE > 15%, Market Cap > 500 Cr")
summary_lines.append("- 5Y Sales CAGR > 10%, 5Y PAT CAGR > 12%")
summary_lines.append("- OPM > 15% (5Y avg), CFO/OP > 70% (5Y avg)")
summary_lines.append("- Debtor days < 90, D/E < 1, CFO positive 4+ of 5 years")
summary_lines.append("")
summary_lines.append("---")
summary_lines.append("")

# Summary table
summary_lines.append("## Ranked Summary (by Quality Score + Forward CAGR)")
summary_lines.append("")
summary_lines.append("| # | Ticker | Sector | Score | CMP | MCap(Cr) | ROCE | OPM | CFO/OP | 5Y-Base | 10Y-Base | Verdict |")
summary_lines.append("|---|--------|--------|-------|-----|----------|------|-----|--------|---------|----------|---------|")

for i, r in enumerate(all_results, 1):
    summary_lines.append(
        f"| {i} | **{r['ticker']}** | {r['sector'][:15]} | {r['quality_score']} | {r.get('cmp', '-')} | {r.get('mcap', 0):,.0f} | {r.get('roce', 0):.0f}% | {r.get('opm_avg', 0):.0f}% | {r.get('avg_cfo_op', 0):.0f}% | {r.get('cagr_5y_base', 0):.0f}% | {r.get('cagr_10y_base', 0):.0f}% | {r['verdict']} |"
    )

summary_lines.append("")
summary_lines.append("---")
summary_lines.append("")

# Tier breakdown
for tier, min_score, max_score in [("Tier 1: Highest Quality", 13, 99), ("Tier 2: High Quality", 11, 12), ("Tier 3: Quality", 9, 10), ("Tier 4: Watchlist", 0, 8)]:
    tier_cos = [r for r in all_results if min_score <= r['quality_score'] <= max_score]
    if not tier_cos:
        continue
    summary_lines.append(f"## {tier}")
    summary_lines.append("")
    for r in tier_cos:
        summary_lines.append(f"### {r['name']} ({r['ticker']})")
        summary_lines.append(f"**Score:** {r['quality_score']} | **Sector:** {r['sector']} | **ROCE:** {r.get('roce', 0):.1f}% | **CMP:** Rs {r.get('cmp', '-')}")
        summary_lines.append("")

        # Key strengths
        summary_lines.append("**Strengths:** " + " | ".join(r.get('quality_reasons', [])[:5]))

        # Forward view
        if r.get('cagr_5y_base') is not None:
            summary_lines.append(f"**5Y outlook:** Bear {r.get('cagr_5y_bear', 0):.0f}% / Base {r.get('cagr_5y_base', 0):.0f}% / Bull {r.get('cagr_5y_bull', 0):.0f}% CAGR")
            summary_lines.append(f"**10Y outlook:** Bear {r.get('cagr_10y_bear', 0):.0f}% / Base {r.get('cagr_10y_base', 0):.0f}% / Bull {r.get('cagr_10y_bull', 0):.0f}% CAGR")

        if r.get('incremental_roce'):
            summary_lines.append(f"**Incremental ROCE:** {r['incremental_roce']:.1f}%")

        # Flags
        flags = []
        if (r.get('opm_range') or 0) > 10: flags.append(f"OPM volatile ({r['opm_range']:.0f}% range)")
        if (r.get('debtor_days') or 0) > 60: flags.append(f"High debtor days ({r['debtor_days']:.0f})")
        if (r.get('promoter_change') or 0) < -3: flags.append(f"Promoter selling ({r['promoter_change']:+.1f}pp)")
        if (r.get('de_ratio') or 0) > 0.5: flags.append(f"Moderate debt (D/E {r['de_ratio']:.2f})")
        if r.get('wc_trend') == 'Deteriorating': flags.append("Working capital deteriorating")
        if not r.get('rev_accelerating'): flags.append("Revenue decelerating")

        if flags:
            summary_lines.append(f"**Watch:** {' | '.join(flags)}")

        summary_lines.append("")

summary_lines.append("---")
summary_lines.append(f"*Analysis based on Screener.in data as of Apr 2026. Forward projections use incremental ROCE and historical growth rates.*")

with open(os.path.join(OUTPUT_DIR, "_MASTER_REPORT.md"), 'w', encoding='utf-8') as f:
    f.write("\n".join(summary_lines))

# Also save as JSON for programmatic use
json_data = []
for r in all_results:
    r_copy = {k: v for k, v in r.items() if k != 'quality_reasons'}
    r_copy['quality_reasons'] = r.get('quality_reasons', [])
    json_data.append(r_copy)

with open(os.path.join(OUTPUT_DIR, "_quality_data.json"), 'w', encoding='utf-8') as f:
    json.dump(json_data, f, indent=2, default=str)

print(f"\n{'='*60}")
print(f"DONE. {len(all_results)} reports generated.")
print(f"Individual reports: {OUTPUT_DIR}/<TICKER>.md")
print(f"Master report: {OUTPUT_DIR}/_MASTER_REPORT.md")
print(f"JSON data: {OUTPUT_DIR}/_quality_data.json")
print(f"{'='*60}")

# Print top 15 by base case 10Y CAGR
print(f"\nTOP 15 BY 10-YEAR BASE CASE CAGR:")
print(f"{'Ticker':<16} {'Score':>5} {'ROCE':>6} {'OPM':>5} {'CFO/OP':>7} {'5Y CAGR':>8} {'10Y CAGR':>9} {'Verdict'}")
print("-" * 80)
by_cagr = sorted(all_results, key=lambda x: x.get('cagr_10y_base', 0), reverse=True)
for r in by_cagr[:15]:
    print(f"{r['ticker']:<16} {r['quality_score']:>5} {r.get('roce',0):>5.1f}% {r.get('opm_avg',0):>4.0f}% {r.get('avg_cfo_op',0):>6.0f}% {r.get('cagr_5y_base',0):>7.1f}% {r.get('cagr_10y_base',0):>8.1f}% {r['verdict']}")
