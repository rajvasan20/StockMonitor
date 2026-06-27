import sys, io, os, warnings
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
warnings.filterwarnings('ignore')
import openpyxl
import statistics

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "ticker_excels")
files = [f for f in os.listdir(DATA_DIR) if f.endswith('.xlsx')]
print(f"Scanning {len(files)} companies (MCap < 1500 Cr filter)...\n")

results = []

def safe_float(v):
    if v is None or v == 'N/A' or v == '':
        return None
    try:
        if isinstance(v, str):
            v = v.replace(',', '').replace('%', '').strip()
        return float(v)
    except:
        return None

def get_val(ws, key):
    for row in ws.iter_rows(values_only=True):
        if row[0] and str(row[0]).strip() == key:
            return row[1]
    return None

for fname in files:
    ticker = fname.replace('.xlsx', '')
    try:
        wb = openpyxl.load_workbook(os.path.join(DATA_DIR, fname), data_only=True, read_only=True)

        ws_val = wb['Valuation']
        roce = safe_float(get_val(ws_val, 'ROCE %'))
        roe = safe_float(get_val(ws_val, 'ROE %'))
        mcap = safe_float(get_val(ws_val, 'Market Cap (Cr)'))
        pe = safe_float(get_val(ws_val, 'Stock P/E'))
        div_yield = safe_float(get_val(ws_val, 'Dividend Yield %'))
        sector = get_val(ws_val, 'Sector')
        cmp = safe_float(get_val(ws_val, 'CMP'))

        if roce is None or roe is None or mcap is None:
            wb.close(); continue
        if mcap >= 1500 or mcap < 100:
            wb.close(); continue
        if roce < 15 or roe < 12:
            wb.close(); continue

        # P&L
        ws_pl = wb['Profit & Loss']
        pl_data = {}
        for row in ws_pl.iter_rows(values_only=True):
            if row[0]:
                pl_data[str(row[0]).strip()] = list(row[1:])
        headers_pl = list(wb['Profit & Loss'].iter_rows(min_row=1, max_row=1, values_only=True))[0][1:]
        sales = [safe_float(x) for x in pl_data.get('Sales', [])]
        opm = [safe_float(x) for x in pl_data.get('OPM %', [])]
        op = [safe_float(x) for x in pl_data.get('Operating Profit', [])]
        np_row = [safe_float(x) for x in pl_data.get('Net Profit', [])]

        year_cols = [str(h) for h in headers_pl if h and 'TTM' not in str(h)]
        n_years = len(year_cols)
        if n_years < 5:
            wb.close(); continue

        sales_clean = [s for s in sales[:n_years] if s is not None]
        opm_clean = [o for o in opm[:n_years] if o is not None]
        op_clean = [o for o in op[:n_years] if o is not None]
        np_clean = [n for n in np_row[:n_years] if n is not None]

        if len(sales_clean) < 5 or len(opm_clean) < 4 or len(np_clean) < 5:
            wb.close(); continue

        s_latest = sales_clean[-1]
        yrs = min(5, len(sales_clean) - 1)
        s_base = sales_clean[-(yrs + 1)]
        if s_base <= 0 or s_latest <= 0:
            wb.close(); continue
        sales_cagr_5y = ((s_latest / s_base) ** (1 / yrs) - 1) * 100

        s_3y = sales_clean[-4] if len(sales_clean) >= 4 else sales_clean[0]
        sales_cagr_3y = ((s_latest / s_3y) ** (1 / 3) - 1) * 100 if s_3y > 0 else 0

        if sales_cagr_5y < 8:
            wb.close(); continue

        np_latest = np_clean[-1]
        np_base = np_clean[-(yrs + 1)] if len(np_clean) > yrs else np_clean[0]
        if np_base <= 0 or np_latest <= 0:
            wb.close(); continue
        pat_cagr_5y = ((np_latest / np_base) ** (1 / yrs) - 1) * 100

        if pat_cagr_5y < 10:
            wb.close(); continue

        opm_last5 = [o for o in opm_clean[-5:] if o is not None]
        if len(opm_last5) < 3:
            wb.close(); continue
        opm_range = max(opm_last5) - min(opm_last5)
        opm_avg = statistics.mean(opm_last5)
        if opm_avg < 12:
            wb.close(); continue

        # Cash Flow
        ws_cf = wb['Cash Flow']
        cf_data = {}
        for row in ws_cf.iter_rows(values_only=True):
            if row[0]:
                cf_data[str(row[0]).strip()] = list(row[1:])
        cfo = [safe_float(x) for x in cf_data.get('Cash from Operating Activity', [])]
        cfo_clean = [c for c in cfo if c is not None]
        cfo_last5 = cfo_clean[-5:] if len(cfo_clean) >= 5 else cfo_clean
        op_last5 = op_clean[-5:] if len(op_clean) >= 5 else op_clean

        if len(cfo_last5) < 3 or len(op_last5) < 3:
            wb.close(); continue

        min_len = min(len(cfo_last5), len(op_last5))
        cfo_op_ratios = []
        cfo_positive_years = 0
        for i in range(min_len):
            if op_last5[i] and op_last5[i] > 0:
                ratio = cfo_last5[i] / op_last5[i] * 100
                cfo_op_ratios.append(ratio)
                if cfo_last5[i] > 0:
                    cfo_positive_years += 1

        if len(cfo_op_ratios) < 2:
            wb.close(); continue
        avg_cfo_op = statistics.mean(cfo_op_ratios)
        if avg_cfo_op < 60:
            wb.close(); continue

        # Ratios
        ws_rat = wb['Ratios']
        rat_data = {}
        for row in ws_rat.iter_rows(values_only=True):
            if row[0]:
                rat_data[str(row[0]).strip()] = list(row[1:])
        debtor_days = [safe_float(x) for x in rat_data.get('Debtor Days', [])]
        dd_clean = [d for d in debtor_days if d is not None]
        latest_dd = dd_clean[-1] if dd_clean else 999
        if latest_dd > 120:
            wb.close(); continue

        # Shareholding
        ws_sh = wb['Shareholding']
        sh_data = {}
        for row in ws_sh.iter_rows(values_only=True):
            if row[0]:
                sh_data[str(row[0]).strip()] = list(row[1:])
        promoter = [safe_float(x) for x in sh_data.get('Promoters', [])]
        prom_clean = [p for p in promoter if p is not None]
        prom_latest = prom_clean[-1] if prom_clean else 0
        prom_earliest = prom_clean[0] if prom_clean else 0
        prom_change = prom_latest - prom_earliest if prom_latest and prom_earliest else 0

        # Balance Sheet
        ws_bs = wb['Balance Sheet']
        bs_data = {}
        for row in ws_bs.iter_rows(values_only=True):
            if row[0]:
                bs_data[str(row[0]).strip()] = list(row[1:])
        borrowings = [safe_float(x) for x in bs_data.get('Borrowings', [])]
        reserves = [safe_float(x) for x in bs_data.get('Reserves', [])]
        equity = [safe_float(x) for x in bs_data.get('Equity Capital', [])]
        borr_latest = [b for b in borrowings if b is not None][-1] if any(b is not None for b in borrowings) else 0
        res_latest = [r for r in reserves if r is not None][-1] if any(r is not None for r in reserves) else 1
        eq_latest = [e for e in equity if e is not None][-1] if any(e is not None for e in equity) else 1
        total_equity = (res_latest or 0) + (eq_latest or 0)
        de_ratio = (borr_latest or 0) / total_equity if total_equity > 0 else 99
        if de_ratio > 1.5:
            wb.close(); continue

        rev_accelerating = sales_cagr_3y >= (sales_cagr_5y * 0.7)
        op_leverage = pat_cagr_5y > sales_cagr_5y

        # SCORING
        score = 0
        if opm_range <= 5: score += 3
        elif opm_range <= 8: score += 2
        elif opm_range <= 12: score += 1
        if avg_cfo_op >= 100: score += 3
        elif avg_cfo_op >= 85: score += 2
        elif avg_cfo_op >= 70: score += 1
        if sales_cagr_5y >= 18: score += 2
        elif sales_cagr_5y >= 12: score += 1
        if pat_cagr_5y >= 20: score += 2
        elif pat_cagr_5y >= 15: score += 1
        if roce >= 25: score += 2
        elif roce >= 20: score += 1
        if rev_accelerating: score += 1
        if op_leverage: score += 1
        if prom_change > 0: score += 1
        if de_ratio < 0.1: score += 1
        if latest_dd < 30: score += 1

        results.append({
            'ticker': ticker, 'sector': (sector or '-')[:17], 'mcap': mcap,
            'roce': roce, 'roe': roe, 'pe': pe or 0,
            'sales_5y': round(sales_cagr_5y, 1), 'sales_3y': round(sales_cagr_3y, 1),
            'pat_5y': round(pat_cagr_5y, 1), 'opm_avg': round(opm_avg, 1),
            'opm_range': round(opm_range, 1), 'cfo_op': round(avg_cfo_op, 1),
            'dd': round(latest_dd), 'de': round(de_ratio, 2),
            'prom': round(prom_latest, 1) if prom_latest else 0,
            'prom_chg': round(prom_change, 1), 'score': score,
            'accel': 'Y' if rev_accelerating else 'N',
            'leverage': 'Y' if op_leverage else 'N',
            'div': div_yield or 0,
        })
        wb.close()
    except Exception as e:
        continue

results.sort(key=lambda x: x['score'], reverse=True)

print(f"Companies passing quality filters (MCap < 1500 Cr): {len(results)}\n")
hdr = f"{'Ticker':<16} {'Sector':<18} {'MCap':>7} {'PE':>6} {'ROCE':>5} {'ROE':>5} {'Sal5Y':>6} {'Sal3Y':>6} {'PAT5Y':>6} {'OPM':>5} {'Range':>5} {'CFO/OP':>6} {'DD':>4} {'D/E':>5} {'Prom%':>6} {'PChg':>5} {'Scr':>4} {'Acc':>3} {'OpL':>3}"
print(hdr)
print("-" * len(hdr))
for r in results:
    print(f"{r['ticker']:<16} {r['sector']:<18} {r['mcap']:>7.0f} {r['pe']:>6.1f} {r['roce']:>5.1f} {r['roe']:>5.1f} {r['sales_5y']:>6.1f} {r['sales_3y']:>6.1f} {r['pat_5y']:>6.1f} {r['opm_avg']:>5.1f} {r['opm_range']:>5.1f} {r['cfo_op']:>6.1f} {r['dd']:>4} {r['de']:>5.2f} {r['prom']:>6.1f} {r['prom_chg']:>5.1f} {r['score']:>4} {r['accel']:>3} {r['leverage']:>3}")
