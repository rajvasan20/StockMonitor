"""
Deep Analysis Agent
-------------------
Runs /past-performance and /valuation skills for quality companies via Claude API.
Produces consultant-grade analysis with year-by-year tables, narratives, and verdicts.

Usage:
    python run.py analyze TRITURBINE              # single company
    python run.py analyze-batch                   # all 43 quality companies
    python run.py analyze TRITURBINE --skill past-performance  # specific skill only
    python run.py analyze TRITURBINE --skill valuation
    python run.py analyze-batch --dry-run          # show what would be processed

Requirements:
    pip install anthropic openpyxl
    Set ANTHROPIC_API_KEY in environment or .env file
"""

import os
import sys
import json
import base64
import time
import re
import io
from datetime import datetime
from pathlib import Path

import anthropic
import openpyxl

from config import (
    BASE_DIR, TICKER_EXCELS_DIR, ANNUAL_REPORTS_DIR,
    QUALITY_DATA_PATH,
)

# ── Paths ────────────────────────────────────────────────────────────────────
SKILL_DIR = Path(os.path.expanduser("~")) / ".claude" / "commands"
PAST_PERF_SKILL = SKILL_DIR / "past-performance.md"
VALUATION_SKILL = SKILL_DIR / "valuation.md"

PAST_PERF_OUTPUT = Path(BASE_DIR) / "output" / "past_performance"
VALUATION_OUTPUT = Path(BASE_DIR) / "output" / "valuation"
LOG_FILE = Path(BASE_DIR) / "scripts" / "_analysis_agent.log"

MODEL = "claude-sonnet-4-6"
MAX_TOKENS = 12000
DELAY_SECONDS = 3


def log(msg, also_print=True):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{timestamp}] {msg}"
    if also_print:
        print(line)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line + "\n")


def load_skill(skill_path):
    if not skill_path.exists():
        raise FileNotFoundError(f"Skill file not found: {skill_path}")
    return skill_path.read_text(encoding="utf-8")


def get_quality_tickers():
    """Load list of quality tickers from _quality_data.json."""
    with open(QUALITY_DATA_PATH, encoding="utf-8") as f:
        data = json.load(f)
    return [d["ticker"] for d in data]


def extract_excel_data(ticker):
    """Read Excel file and extract all sheets as formatted text for the API.

    Returns a string representation of all sheets that Claude can analyze.
    """
    excel_path = Path(TICKER_EXCELS_DIR) / f"{ticker}.xlsx"
    if not excel_path.exists():
        return None

    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    sections = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            row_vals = []
            for cell in row:
                if cell is None:
                    row_vals.append("")
                else:
                    row_vals.append(str(cell))
            rows.append(" | ".join(row_vals))

        if rows:
            sections.append(f"### Sheet: {sheet_name}\n")
            # Header row
            sections.append(rows[0])
            sections.append("-" * 40)
            for r in rows[1:]:
                sections.append(r)
            sections.append("")

    wb.close()
    return "\n".join(sections)


def find_mda_context(ticker):
    """Find and read the latest MDA context files for a ticker."""
    ar_dir = Path(ANNUAL_REPORTS_DIR) / ticker
    if not ar_dir.exists():
        return ""

    mda_files = sorted(ar_dir.glob(f"{ticker}_MDA_*_context.md"), reverse=True)
    if not mda_files:
        return ""

    # Read latest 2-3 MDA files for context
    context_parts = []
    for mda_file in mda_files[:3]:
        content = mda_file.read_text(encoding="utf-8")
        # Truncate to key blocks (1, 1A, 2, 3, 6, 7) to save tokens
        context_parts.append(f"--- {mda_file.name} ---\n{content[:4000]}\n")

    return "\n".join(context_parts)


def run_skill(client, skill_prompt, skill_name, ticker, excel_data, mda_context=""):
    """Send Excel data + skill to Claude API. Returns markdown response."""

    user_content = (
        f"Run the {skill_name} analysis for ticker: {ticker}\n\n"
        f"## Financial Data (from {ticker}.xlsx)\n\n"
        f"{excel_data}\n\n"
    )

    if mda_context:
        user_content += (
            f"## MD&A Context (from annual reports)\n\n"
            f"{mda_context}\n\n"
        )

    user_content += (
        f"Follow all steps in the skill instructions exactly. "
        f"Return only the completed markdown analysis — "
        f"no preamble, no commentary outside the markdown."
    )

    response = client.messages.create(
        model=MODEL,
        max_tokens=MAX_TOKENS,
        system=skill_prompt,
        messages=[
            {
                "role": "user",
                "content": user_content,
            }
        ],
    )

    return response.content[0].text


def process_ticker(client, ticker, skills_to_run, past_perf_prompt, valuation_prompt, dry_run=False):
    """Run specified skills for a single ticker. Returns dict of results."""
    results = {}

    # Load Excel data
    excel_data = extract_excel_data(ticker)
    if not excel_data:
        log(f"  SKIP  {ticker} — no Excel file found")
        return {"status": "skipped", "reason": "no_excel"}

    # Load MDA context (optional)
    mda_context = find_mda_context(ticker)

    if dry_run:
        log(f"  DRY   {ticker} — skills: {', '.join(skills_to_run)}")
        return {"status": "dry-run"}

    # Run past-performance
    if "past-performance" in skills_to_run:
        pp_output = PAST_PERF_OUTPUT / f"{ticker}_past_performance.md"
        if pp_output.exists():
            log(f"  SKIP  {ticker} past-performance — already exists")
            results["past-performance"] = "skipped"
        else:
            log(f"  PROC  {ticker} past-performance")
            try:
                markdown = run_skill(
                    client, past_perf_prompt, "past-performance",
                    ticker, excel_data, mda_context
                )
                PAST_PERF_OUTPUT.mkdir(parents=True, exist_ok=True)
                pp_output.write_text(markdown, encoding="utf-8")
                log(f"  DONE  {ticker} past-performance ({len(markdown)} chars)")
                results["past-performance"] = "done"
            except Exception as e:
                log(f"  FAIL  {ticker} past-performance — {e}")
                results["past-performance"] = f"failed: {e}"

    # Run valuation
    if "valuation" in skills_to_run:
        val_output = VALUATION_OUTPUT / f"{ticker}_valuation.md"
        if val_output.exists():
            log(f"  SKIP  {ticker} valuation — already exists")
            results["valuation"] = "skipped"
        else:
            log(f"  PROC  {ticker} valuation")
            try:
                markdown = run_skill(
                    client, valuation_prompt, "valuation",
                    ticker, excel_data, mda_context
                )
                VALUATION_OUTPUT.mkdir(parents=True, exist_ok=True)
                val_output.write_text(markdown, encoding="utf-8")
                log(f"  DONE  {ticker} valuation ({len(markdown)} chars)")
                results["valuation"] = "done"
            except Exception as e:
                log(f"  FAIL  {ticker} valuation — {e}")
                results["valuation"] = f"failed: {e}"

    return results


def run(ticker=None, skill=None, dry_run=False, force=False):
    """Main entry point.

    Args:
        ticker: Single ticker or None for batch (all quality companies)
        skill: 'past-performance', 'valuation', or None (both)
        dry_run: Show what would be processed without running
        force: Re-run even if output already exists
    """
    # API key
    api_key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if not api_key and not dry_run:
        env_file = Path(BASE_DIR) / ".env"
        if env_file.exists():
            for line in env_file.read_text().splitlines():
                if line.startswith("ANTHROPIC_API_KEY="):
                    api_key = line.split("=", 1)[1].strip().strip('"').strip("'")
                    break
    if not api_key and not dry_run:
        print("ERROR: ANTHROPIC_API_KEY not set.")
        return

    # Determine skills to run
    skills_to_run = []
    if skill:
        if skill in ("past-performance", "valuation"):
            skills_to_run = [skill]
        else:
            print(f"ERROR: Unknown skill '{skill}'. Use 'past-performance' or 'valuation'.")
            return
    else:
        skills_to_run = ["past-performance", "valuation"]

    # Load skill prompts
    past_perf_prompt = load_skill(PAST_PERF_SKILL)
    valuation_prompt = load_skill(VALUATION_SKILL)

    # Client
    client = anthropic.Anthropic(api_key=api_key) if not dry_run else None

    # Determine tickers
    if ticker:
        tickers = [ticker.upper()]
    else:
        tickers = get_quality_tickers()

    # If force, remove existing outputs
    if force and not dry_run:
        for t in tickers:
            for s in skills_to_run:
                if s == "past-performance":
                    out = PAST_PERF_OUTPUT / f"{t}_past_performance.md"
                else:
                    out = VALUATION_OUTPUT / f"{t}_valuation.md"
                if out.exists():
                    out.unlink()
                    log(f"  DEL   {out.name} (force re-run)")

    log(f"{'=' * 60}")
    log(f"Analysis Agent | Tickers: {len(tickers)} | Skills: {', '.join(skills_to_run)} | dry-run: {dry_run}")
    log(f"{'=' * 60}")

    counts = {"done": 0, "skipped": 0, "failed": 0}

    for i, t in enumerate(tickers, 1):
        log(f"[{i}/{len(tickers)}] {t}")
        results = process_ticker(
            client, t, skills_to_run,
            past_perf_prompt, valuation_prompt,
            dry_run=dry_run
        )

        for s, status in results.items():
            if status == "done":
                counts["done"] += 1
            elif status == "skipped":
                counts["skipped"] += 1
            elif isinstance(status, str) and status.startswith("failed"):
                counts["failed"] += 1

        # Rate limit delay
        if not dry_run and any(v == "done" for v in results.values()) and i < len(tickers):
            time.sleep(DELAY_SECONDS)

    log(f"{'=' * 60}")
    log(f"Complete | done={counts['done']} skipped={counts['skipped']} failed={counts['failed']}")
    log(f"{'=' * 60}")
